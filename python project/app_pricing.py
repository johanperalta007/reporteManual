"""
App Reporte Pricing - Interfaz Gráfica
Descarga archivos desde MFT (RPA) y genera el reporte de Pricing.
Muestra el progreso en tiempo real en una consola integrada.

Requisitos:
    pip3 install pandas openpyxl customtkinter selenium webdriver-manager
"""

import os
import sys
import glob
import time
import shutil
import zipfile
import threading
from datetime import datetime
import urllib3

# Deshabilitar verificación SSL (necesario por proxy corporativo)
os.environ['WDM_SSL_VERIFY'] = '0'
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

import customtkinter as ctk
from tkinter import filedialog, messagebox
import configparser

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

# --- Configuración desde archivo ---
def obtener_carpeta_config():
    """Retorna la carpeta base de configuración según el entorno."""
    if getattr(sys, 'frozen', False):
        return os.path.join(os.path.expanduser("~"), "Documents", "ReportePricing")
    else:
        return os.path.dirname(os.path.abspath(__file__))

def cargar_config():
    """Carga la configuración desde config.ini. Si no existe, lo crea con valores por defecto."""
    carpeta = obtener_carpeta_config()
    os.makedirs(carpeta, exist_ok=True)
    ruta_config = os.path.join(carpeta, "config.ini")

    config = configparser.ConfigParser()

    if not os.path.exists(ruta_config):
        # Crear config.ini con valores por defecto
        config["MFT"] = {
            "usuario": "",
            "password": "",
        }
        with open(ruta_config, "w", encoding="utf-8") as f:
            f.write("# ============================================================\n")
            f.write("# CONFIGURACIÓN DE REPORTE PRICING\n")
            f.write("# Editar este archivo con tus credenciales y rutas.\n")
            f.write("# ============================================================\n\n")
            config.write(f)
        print(f"📄 Archivo de configuración creado en: {ruta_config}")
        print("   ⚠️  Edita config.ini con tus credenciales antes de usar el RPA.")

    config.read(ruta_config, encoding="utf-8")
    return config

CONFIG = cargar_config()

# --- Configuración RPA (desde config.ini) ---
URL_MFT = "https://mft.bancodebogota.com.co:4443/webclient/Login.xhtml"
USUARIO = CONFIG.get("MFT", "usuario", fallback="")
PASSWORD = CONFIG.get("MFT", "password", fallback="")
CARPETA_DESCARGA = os.path.join(os.path.expanduser("~"), "Downloads")
TIMEOUT_SELENIUM = 30


# --- Redirigir prints a la consola de la GUI ---
class ConsoleRedirector:
    def __init__(self, textbox):
        self.textbox = textbox

    def write(self, message):
        if message.strip() == "":
            return
        self.textbox.after(0, self._append, message)

    def _append(self, message):
        self.textbox.configure(state="normal")
        self.textbox.insert("end", message + "\n")
        self.textbox.see("end")
        self.textbox.configure(state="disabled")

    def flush(self):
        pass


# --- Lógica RPA: Descarga desde MFT ---
def ejecutar_rpa(carpeta_destino):
    """Ejecuta el RPA completo: login, descarga, descomprime y copia a desembolsos_diarios."""
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from webdriver_manager.chrome import ChromeDriverManager

    def esperar_elemento(driver, by, valor, timeout=TIMEOUT_SELENIUM):
        return WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((by, valor))
        )

    driver = None
    try:
        # Validar credenciales
        if not USUARIO or not PASSWORD:
            carpeta = obtener_carpeta_config()
            ruta_config = os.path.join(carpeta, "config.ini")
            print(f"❌ Credenciales no configuradas.")
            print(f"   Edita el archivo: {ruta_config}")
            print(f"   Agrega tu usuario y contraseña en la sección [MFT]")
            return False

        # Crear driver
        print("🚀 Iniciando RPA - Descarga desde MFT...")
        chrome_options = Options()
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--ignore-certificate-errors")
        chrome_options.add_argument("--allow-insecure-localhost")
        chrome_options.add_experimental_option("prefs", {
            "download.default_directory": CARPETA_DESCARGA,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
        })
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.maximize_window()

        # Login
        print("🌐 Navegando a MFT...")
        driver.get(URL_MFT)
        time.sleep(2)

        print("👤 Ingresando usuario...")
        campo_usuario = esperar_elemento(driver, By.ID, "username")
        campo_usuario.clear()
        campo_usuario.send_keys(USUARIO)

        print("🔑 Ingresando contraseña...")
        campo_password = esperar_elemento(driver, By.ID, "value")
        campo_password.clear()
        campo_password.send_keys(PASSWORD)

        print("🔓 Iniciando sesión...")
        boton_login = esperar_elemento(driver, By.ID, "j_id_1o")
        boton_login.click()
        time.sleep(3)
        print("✅ Login exitoso.")

        # Navegar a carpeta
        print("📂 Navegando a la carpeta de archivos...")
        enlace_carpeta = esperar_elemento(driver, By.ID, "fileListForm:j_id_8l:0:j_id_8s")
        enlace_carpeta.click()
        time.sleep(3)
        print("✅ Carpeta abierta.")

        # Seleccionar todos
        print("☑️  Seleccionando todos los archivos...")
        try:
            checkbox = WebDriverWait(driver, TIMEOUT_SELENIUM).until(
                EC.element_to_be_clickable((
                    By.CSS_SELECTOR,
                    "div.ui-chkbox-box.ui-widget.ui-corner-all.ui-state-default"
                ))
            )
            checkbox.click()
        except Exception:
            print("⚠️  Click directo falló, intentando con JavaScript...")
            checkbox = driver.find_element(
                By.CSS_SELECTOR,
                "div.ui-chkbox-box.ui-widget.ui-corner-all.ui-state-default"
            )
            driver.execute_script("arguments[0].click();", checkbox)
        time.sleep(2)
        print("✅ Todos los archivos seleccionados.")

        # Descargar
        print("⬇️  Descargando archivos...")
        boton_descarga = esperar_elemento(driver, By.ID, "downloadFiles:downloadFiles")
        boton_descarga.click()
        time.sleep(5)
        print("✅ Descarga iniciada.")

        # Esperar descargas
        print("⏳ Esperando a que finalicen las descargas...")
        inicio = time.time()
        while time.time() - inicio < 120:
            archivos = os.listdir(CARPETA_DESCARGA)
            if archivos and not any(f.endswith(".crdownload") for f in archivos):
                print("✅ Descarga completada.")
                break
            time.sleep(2)

        # Buscar ZIP más reciente
        print("🔍 Buscando el ZIP más reciente en Downloads...")
        archivos_zip = glob.glob(os.path.join(CARPETA_DESCARGA, "*.zip"))
        if not archivos_zip:
            raise FileNotFoundError("No se encontró ningún archivo .zip en Downloads")
        zip_reciente = max(archivos_zip, key=os.path.getmtime)
        print(f"✅ ZIP encontrado: {os.path.basename(zip_reciente)}")

        # Descomprimir y copiar
        carpeta_temp = os.path.join(CARPETA_DESCARGA, "_temp_rpa_extract")
        if os.path.exists(carpeta_temp):
            shutil.rmtree(carpeta_temp)

        print(f"📦 Descomprimiendo {os.path.basename(zip_reciente)}...")
        with zipfile.ZipFile(zip_reciente, 'r') as zip_ref:
            zip_ref.extractall(carpeta_temp)

        contenido = os.listdir(carpeta_temp)
        if len(contenido) == 1 and os.path.isdir(os.path.join(carpeta_temp, contenido[0])):
            carpeta_origen = os.path.join(carpeta_temp, contenido[0])
        else:
            carpeta_origen = carpeta_temp

        os.makedirs(carpeta_destino, exist_ok=True)
        reemplazados = 0
        nuevos = 0
        for item in os.listdir(carpeta_origen):
            origen = os.path.join(carpeta_origen, item)
            destino = os.path.join(carpeta_destino, item)
            ya_existia = os.path.exists(destino)
            if os.path.isfile(origen):
                shutil.copy2(origen, destino)
            elif os.path.isdir(origen):
                if os.path.exists(destino):
                    shutil.rmtree(destino)
                shutil.copytree(origen, destino)
            if ya_existia:
                reemplazados += 1
            else:
                nuevos += 1

        print(f"📋 Archivos copiados a desembolsos_diarios/")
        print(f"   📄 Reemplazados: {reemplazados} | Nuevos: {nuevos}")

        shutil.rmtree(carpeta_temp)
        print("🧹 Carpeta temporal eliminada.")
        print("✅ RPA completado exitosamente.\n")
        return True

    except Exception as e:
        print(f"❌ Error en RPA: {e}")
        if driver:
            driver.save_screenshot("error_screenshot.png")
            print("📸 Screenshot de error guardado.")
        return False
    finally:
        if driver:
            driver.quit()
            print("🔒 Navegador cerrado.")


# --- Lógica de procesamiento ---
def ejecutar_procesamiento(archivo_pricing, archivo_suite, carpeta_txt, salida_final):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Color
    from copy import copy

    COL_SPREAD_REAL = "AT"
    COL_SPREAD_SUGERIDO = "AU"
    COL_DESVIACION = "AV"
    COLOR_VERDE = "FF00B050"
    COLOR_ROJO = "FFC00000"

    def normalizar_numero_credito(valor):
        if pd.isna(valor) or str(valor).strip() == '':
            return valor
        numero = str(valor).strip()
        if '.' in numero:
            numero = numero.split('.')[0]
        if len(numero) >= 14:
            return numero
        return numero.zfill(14)

    def _to_float(v):
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip()
        if s == "":
            return None
        s = s.replace(",", ".")
        try:
            return float(s)
        except:
            return None

    try:
        print("📂 Cargando archivo de Pricing...")
        df_pricing = pd.read_excel(archivo_pricing)
        print(f"   ✅ {len(df_pricing)} registros cargados desde Pricing")

        print("📂 Cargando archivo de Suite Digital...")
        df_suite = pd.read_excel(archivo_suite)
        print(f"   ✅ {len(df_suite)} registros cargados desde Suite Digital")

        print(f"📄 Procesando archivos TXT de desembolsos...")
        registros = []
        archivos_txt = [f for f in os.listdir(carpeta_txt) if f.endswith(".txt")]
        print(f"   📁 {len(archivos_txt)} archivos TXT encontrados")

        for archivo in archivos_txt:
            ruta = os.path.join(carpeta_txt, archivo)
            with open(ruta, 'r', encoding='utf-8') as f:
                lineas = f.readlines()
                for linea in lineas[1:]:
                    partes = linea.strip().split('|')
                    if len(partes) >= 3:
                        cotizacion = partes[0].replace('SOL_PRC_', '')
                        numero_credito = partes[1]
                        fecha_raw = partes[2]
                        try:
                            fecha = datetime.strptime(fecha_raw, '%Y%m%d').date()
                        except:
                            fecha = None
                        registros.append({
                            'Cotizacion_Num': cotizacion,
                            'Número de Crédito 1': numero_credito,
                            'Fecha de Desembolso 1': fecha
                        })
        df_txt = pd.DataFrame(registros)
        print(f"   ✅ {len(df_txt)} registros de desembolsos procesados")

        print("🔧 Limpiando cotizaciones...")
        df_pricing['Cotizacion_Num'] = df_pricing['Número Cotización'].astype(str)
        df_pricing['Cotizacion_Num'] = df_pricing['Cotizacion_Num'].str.replace('SOL_PRC_', '', regex=False)
        df_pricing['Cotizacion_Num'] = df_pricing['Cotizacion_Num'].str.replace('PRI', '', regex=False)

        print("🔗 Cruzando datos con Suite Digital...")
        df_suite = df_suite.rename(columns={'x': 'Cotizacion_Num'})
        df_suite['Cotizacion_Num'] = df_suite['Cotizacion_Num'].astype(str)
        df_pricing['Cotizacion_Num'] = df_pricing['Cotizacion_Num'].astype(str)

        df_suite_min = df_suite[['Cotizacion_Num', 'NUMERO_CREDITO', 'FECHA_ACTUALIZACION']].drop_duplicates()
        df_merge = pd.merge(df_pricing, df_suite_min, on='Cotizacion_Num', how='left')
        df_merge['Número de Crédito 1'] = df_merge['Número de Crédito 1'].fillna(df_merge['NUMERO_CREDITO'])
        df_merge['Fecha de Desembolso 1'] = df_merge['Fecha de Desembolso 1'].fillna(df_merge['FECHA_ACTUALIZACION'])

        print("🔗 Cruzando datos con archivos TXT...")
        df_txt['Cotizacion_Num'] = df_txt['Cotizacion_Num'].astype(str)
        df_txt_min = df_txt[['Cotizacion_Num', 'Número de Crédito 1', 'Fecha de Desembolso 1']].drop_duplicates()
        df_final = pd.merge(
            df_merge.drop(columns=['NUMERO_CREDITO', 'FECHA_ACTUALIZACION'], errors='ignore'),
            df_txt_min, on='Cotizacion_Num', how='left', suffixes=('', '_txt')
        )
        df_final['Número de Crédito 1'] = df_final['Número de Crédito 1'].fillna(df_final['Número de Crédito 1_txt'])
        df_final['Fecha de Desembolso 1'] = df_final['Fecha de Desembolso 1'].fillna(df_final['Fecha de Desembolso 1_txt'])

        print("🔥 Aplicando filtros...")
        df_final = df_final[~df_final['Radicador'].str.upper().eq('MARIZA@BANCODEBOGOTA.COM.CO')]
        df_final = df_final[~df_final['Estado'].str.upper().eq('BORRADOR')]

        print("🔥 Resolviendo duplicados (priorizando aprobados)...")
        prioridad_estado = {"APROBADO": 1, "APROBADA": 1, "APROBADA DIGITAL": 1, "APROBADA WEB": 1}
        df_final["prioridad"] = df_final["Estado"].str.upper().map(prioridad_estado).fillna(2)
        df_final = df_final.sort_values(by=["Número Cotización", "prioridad"], ascending=[True, True])
        df_final = df_final.drop_duplicates(subset=["Número Cotización"], keep="first")
        df_final['Número de Crédito 1'] = df_final['Número de Crédito 1'].apply(normalizar_numero_credito)
        df_final = df_final.sort_values(by="Número Cotización", ascending=True)
        df_final = df_final.drop(columns=[
            'Cotizacion_Num', 'Número de Crédito 1_txt', 'Fecha de Desembolso 1_txt', 'prioridad'
        ], errors='ignore')

        print(f"📊 Registros finales: {len(df_final)}")

        print(f"💾 Guardando reporte en: {os.path.basename(salida_final)}")
        df_final.to_excel(salida_final, index=False)
        print("✅ Archivo Excel guardado")

        print("🎨 Pintando columna Desviación (AV)...")
        wb = load_workbook(salida_final)
        ws = wb.active
        verde = Color(rgb=COLOR_VERDE)
        rojo = Color(rgb=COLOR_ROJO)
        pintadas = 0
        for r in range(2, ws.max_row + 1):
            spread = _to_float(ws[f"{COL_SPREAD_REAL}{r}"].value)
            sugerido = _to_float(ws[f"{COL_SPREAD_SUGERIDO}{r}"].value)
            if spread is None or sugerido is None:
                continue
            celda_av = ws[f"{COL_DESVIACION}{r}"]
            fuente_actual = celda_av.font if celda_av.font else Font()
            nueva = copy(fuente_actual)
            if spread > sugerido:
                nueva.color = verde
                celda_av.font = nueva
                pintadas += 1
            elif spread < sugerido:
                nueva.color = rojo
                celda_av.font = nueva
                pintadas += 1
        wb.save(salida_final)
        print(f"✅ {pintadas} celdas pintadas en columna AV")

        print("")
        print("=" * 50)
        print("✅ PROCESO COMPLETADO EXITOSAMENTE")
        print(f"📁 Archivo: {os.path.basename(salida_final)}")
        print("=" * 50)
        return True

    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        return False


# --- Interfaz Gráfica con CustomTkinter ---
class AppPricing(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Reporte Pricing - Banco de Bogotá")
        self.geometry("720x650")
        self.resizable(False, False)

        # Ícono de la ventana
        script_dir = os.path.dirname(os.path.abspath(__file__))
        if getattr(sys, 'frozen', False):
            icon_base = sys._MEIPASS
        else:
            icon_base = script_dir
        logo_path = os.path.join(icon_base, "images", "logoWhite.png")
        if os.path.exists(logo_path):
            from PIL import Image, ImageTk
            img = Image.open(logo_path).convert("RGBA").resize((32, 32))
            self._icon = ImageTk.PhotoImage(img)
            self.iconphoto(False, self._icon)

        # Rutas fijas
        # Detectar si corre empaquetado (PyInstaller) o desde código fuente
        if getattr(sys, 'frozen', False):
            # Empaquetado: usar carpeta de trabajo en Documents del usuario
            base_dir = os.path.join(os.path.expanduser("~"), "Documents", "ReportePricing")
            os.makedirs(os.path.join(base_dir, "excelFiles"), exist_ok=True)
            os.makedirs(os.path.join(base_dir, "desembolsos_diarios"), exist_ok=True)
            # Para imágenes empaquetadas dentro del .app
            bundle_dir = sys._MEIPASS
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            bundle_dir = base_dir

        self.base_dir = base_dir
        self.archivo_suite = os.path.join(base_dir, "excelFiles", "Reporte de Suite_Digital.xlsx")
        self.carpeta_txt = os.path.join(base_dir, "desembolsos_diarios")
        self.default_salida = os.path.join(base_dir, "excelFiles")

        # Variables
        self.ruta_pricing = ""
        self.ruta_salida = ""

        self._crear_interfaz()

    def _crear_interfaz(self):
        # Título + marca de agua en la misma línea
        row_title = ctk.CTkFrame(self, fg_color="transparent")
        row_title.pack(fill="x", padx=30, pady=(20, 0))
        ctk.CTkLabel(row_title, text="Reporte Pricing",
                     font=ctk.CTkFont(size=24, weight="bold")).pack(side="left")
        ctk.CTkLabel(row_title, text="Powered By ADL - JP",
                     font=ctk.CTkFont(size=10, slant="italic"),
                     text_color="#666666").pack(side="right", pady=(8, 0))

        # Subtítulo
        ctk.CTkLabel(self, text="Banco de Bogotá  —  Generador de reportes",
                     font=ctk.CTkFont(size=13), text_color="gray").pack(padx=30, anchor="w", pady=(2, 18))

        # --- Card 1: Archivo Pricing ---
        card1 = ctk.CTkFrame(self, corner_radius=10)
        card1.pack(fill="x", padx=30, pady=(0, 12))

        ctk.CTkLabel(card1, text="📄  Archivo de Pricing",
                     font=ctk.CTkFont(size=15, weight="bold")).pack(padx=15, pady=(12, 0), anchor="w")
        ctk.CTkLabel(card1, text="Selecciona el archivo Excel (.xlsx) de Pricing que deseas procesar",
                     font=ctk.CTkFont(size=12), text_color="gray").pack(padx=15, anchor="w", pady=(2, 8))

        row1 = ctk.CTkFrame(card1, fg_color="transparent")
        row1.pack(fill="x", padx=15, pady=(0, 12))

        self.entry_pricing = ctk.CTkEntry(row1, placeholder_text="Ningún archivo seleccionado...",
                                          state="readonly", font=ctk.CTkFont(size=12), height=35)
        self.entry_pricing.pack(side="left", fill="x", expand=True)

        ctk.CTkButton(row1, text="Seleccionar archivo", width=160, height=35,
                      command=self._seleccionar_pricing).pack(side="right", padx=(10, 0))

        # --- Card 2: Carpeta de salida ---
        card2 = ctk.CTkFrame(self, corner_radius=10)
        card2.pack(fill="x", padx=30, pady=(0, 12))

        ctk.CTkLabel(card2, text="📁  Carpeta de salida",
                     font=ctk.CTkFont(size=15, weight="bold")).pack(padx=15, pady=(12, 0), anchor="w")
        ctk.CTkLabel(card2, text="Carpeta donde se guardará el reporte generado automáticamente",
                     font=ctk.CTkFont(size=12), text_color="gray").pack(padx=15, anchor="w", pady=(2, 8))

        row2 = ctk.CTkFrame(card2, fg_color="transparent")
        row2.pack(fill="x", padx=15, pady=(0, 12))

        self.entry_salida = ctk.CTkEntry(row2, font=ctk.CTkFont(size=12), height=35,
                                         placeholder_text="Ninguna carpeta seleccionada...",
                                         state="readonly")
        self.entry_salida.pack(side="left", fill="x", expand=True)

        ctk.CTkButton(row2, text="Seleccionar carpeta", width=160, height=35,
                      command=self._seleccionar_carpeta_salida).pack(side="right", padx=(10, 0))

        # --- Checkbox RPA ---
        self.rpa_activado = ctk.BooleanVar(value=False)
        self.chk_rpa = ctk.CTkCheckBox(
            self, text="🌐  Descargar desembolsos desde MFT antes de procesar (RPA)",
            variable=self.rpa_activado, font=ctk.CTkFont(size=13)
        )
        self.chk_rpa.pack(padx=30, pady=(5, 5), anchor="w")

        # --- Botón procesar ---
        self.btn_procesar = ctk.CTkButton(self, text="▶  Procesar Reporte", height=45,
                                          font=ctk.CTkFont(size=16, weight="bold"),
                                          command=self._iniciar_proceso)
        self.btn_procesar.pack(fill="x", padx=30, pady=(10, 12))

        # --- Consola ---
        ctk.CTkLabel(self, text="Consola de progreso:", font=ctk.CTkFont(size=12, weight="bold")).pack(
            anchor="w", padx=30)
        self.consola = ctk.CTkTextbox(self, font=ctk.CTkFont(family="Courier", size=12),
                                      fg_color="#1a1a1a", text_color="#00ff00",
                                      corner_radius=8, state="disabled")
        self.consola.pack(fill="both", expand=True, padx=30, pady=(4, 18))

    def _seleccionar_pricing(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar archivo de Pricing",
            filetypes=[("Excel", "*.xlsx *.xls")]
        )
        if ruta:
            self.ruta_pricing = ruta
            self.entry_pricing.configure(state="normal")
            self.entry_pricing.delete(0, "end")
            self.entry_pricing.insert(0, ruta)
            self.entry_pricing.configure(state="readonly")

    def _seleccionar_carpeta_salida(self):
        ruta = filedialog.askdirectory(title="Seleccionar carpeta de salida")
        if ruta:
            self.ruta_salida = ruta
            self.entry_salida.configure(state="normal")
            self.entry_salida.delete(0, "end")
            self.entry_salida.insert(0, ruta)
            self.entry_salida.configure(state="readonly")

    def _iniciar_proceso(self):
        if not self.ruta_pricing:
            messagebox.showwarning("Campo requerido", "Por favor selecciona el archivo de Pricing desde tu equipo.")
            return
        if not self.ruta_salida:
            messagebox.showwarning("Campo requerido", "Por favor selecciona una carpeta de salida en tu equipo.")
            return

        # Limpiar consola
        self.consola.configure(state="normal")
        self.consola.delete("1.0", "end")
        self.consola.configure(state="disabled")

        self.btn_procesar.configure(state="disabled", text="⏳ Procesando...")

        sys.stdout = ConsoleRedirector(self.consola)
        sys.stderr = ConsoleRedirector(self.consola)

        hoy = datetime.now().strftime("%d-%m-%Y")
        salida = os.path.join(self.ruta_salida, f"Reporte_Pricing_Actualizado_{hoy}.xlsx")

        def proceso():
            exito = True

            # Paso 1: RPA si está activado
            if self.rpa_activado.get():
                exito = ejecutar_rpa(carpeta_destino=self.carpeta_txt)
                if not exito:
                    self.after(0, self._proceso_terminado, False)
                    return

            # Paso 2: Procesar reporte
            print("📊 Iniciando procesamiento del reporte...")
            resultado = ejecutar_procesamiento(
                archivo_pricing=self.ruta_pricing,
                archivo_suite=self.archivo_suite,
                carpeta_txt=self.carpeta_txt,
                salida_final=salida
            )
            self.after(0, self._proceso_terminado, resultado)

        threading.Thread(target=proceso, daemon=True).start()

    def _proceso_terminado(self, exito):
        sys.stdout = sys.__stdout__
        sys.stderr = sys.__stderr__
        self.btn_procesar.configure(state="normal", text="▶  Procesar Reporte")
        if exito:
            messagebox.showinfo("Listo", "El reporte se generó exitosamente.")


def main():
    app = AppPricing()
    app.mainloop()


if __name__ == "__main__":
    main()

"""
   $$$$$\ $$$$$$$\        $$$$$$$\   $$$$$$\ $$$$$$$$\  $$$$$$\         $$$$$$\ $$\     $$\  $$$$$$\ $$$$$$$$\ $$$$$$$$\ $$\      $$\  $$$$$$\  
   \__$$ |$$  __$$\       $$  __$$\ $$  __$$\\__$$  __|$$  __$$\       $$  __$$\\$$\   $$  |$$  __$$\\__$$  __|$$  _____|$$$\    $$$ |$$  __$$\ 
      $$ |$$ |  $$ |      $$ |  $$ |$$ /  $$ |  $$ |   $$ /  $$ |      $$ /  \__|\$$\ $$  / $$ /  \__|  $$ |   $$ |      $$$$\  $$$$ |$$ /  \__|
      $$ |$$$$$$$  |      $$ |  $$ |$$$$$$$$ |  $$ |   $$$$$$$$ |      \$$$$$$\   \$$$$  /  \$$$$$$\    $$ |   $$$$$\    $$\$$\$$ $$ |\$$$$$$\  
$$\   $$ |$$  ____/       $$ |  $$ |$$  __$$ |  $$ |   $$  __$$ |       \____$$\   \$$  /    \____$$\   $$ |   $$  __|   $$ \$$$  $$ | \____$$\ 
$$ |  $$ |$$ |            $$ |  $$ |$$ |  $$ |  $$ |   $$ |  $$ |      $$\   $$ |   $$ |    $$\   $$ |  $$ |   $$ |      $$ |\$  /$$ |$$\   $$ |
\$$$$$$  |$$ |            $$$$$$$  |$$ |  $$ |  $$ |   $$ |  $$ |      \$$$$$$  |   $$ |    \$$$$$$  |  $$ |   $$$$$$$$\ $$ | \_/ $$ |\$$$$$$  |
 \______/ \__|            \_______/ \__|  \__|  \__|   \__|  \__|       \______/    \__|     \______/   \__|   \________|\__|     \__| \______/ 
"""