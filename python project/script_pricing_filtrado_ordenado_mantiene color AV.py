import pandas as pd
import os
from datetime import datetime
from dotenv import load_dotenv

from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from copy import copy

# --- CARGAR CONFIGURACIÓN DESDE .env ---
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))

CARPETA_EXCEL = os.getenv("CARPETA_EXCEL", "excelFiles")
ARCHIVO_PRICING = os.path.join(CARPETA_EXCEL, os.getenv("ARCHIVO_PRICING", ""))
ARCHIVO_SUITE = os.path.join(CARPETA_EXCEL, os.getenv("ARCHIVO_SUITE", "Reporte de Suite_Digital.xlsx"))
CARPETA_TXT = os.getenv("CARPETA_TXT", "desembolsos_diarios")

# Nombre de salida calculado automáticamente con la fecha de hoy
hoy = datetime.now().strftime("%d-%m-%Y")
SALIDA_FINAL = os.path.join(CARPETA_EXCEL, f"Reporte_Pricing_Actualizado_{hoy}.xlsx")

# Columnas Excel para regla de color (fijas, no se modifican)
COL_SPREAD_REAL = "AT"
COL_SPREAD_SUGERIDO = "AU"
COL_DESVIACION = "AV"

# Colores de fuente (ARGB)
COLOR_VERDE = "FF00B050"
COLOR_ROJO = "FFC00000"

def normalizar_numero_credito(valor):
    if pd.isna(valor) or str(valor).strip() == '':
        return valor
    # Eliminar decimales si viene como float (ej: 1057825976.0 -> 1057825976)
    numero = str(valor).strip()
    if '.' in numero:
        numero = numero.split('.')[0]
    if len(numero) >= 14:
        return numero
    # Rellenar con ceros a la izquierda hasta 14
    return numero.zfill(14)

# ------------------------------------------------------
def cargar_archivos_base():
    df_pricing = pd.read_excel(ARCHIVO_PRICING)
    df_suite = pd.read_excel(ARCHIVO_SUITE)
    return df_pricing, df_suite

# ------------------------------------------------------
def limpiar_cotizaciones(df):
    df['Cotizacion_Num'] = df['Número Cotización'].astype(str)
    df['Cotizacion_Num'] = df['Cotizacion_Num'].str.replace('SOL_PRC_', '', regex=False)
    df['Cotizacion_Num'] = df['Cotizacion_Num'].str.replace('PRI', '', regex=False)
    return df

# ------------------------------------------------------
def procesar_archivos_txt(carpeta):
    registros = []
    for archivo in os.listdir(carpeta):
        if archivo.endswith(".txt"):
            ruta = os.path.join(carpeta, archivo)
            with open(ruta, 'r', encoding='utf-8') as f:
                lineas = f.readlines()
                for linea in lineas[1:]:
                    partes = linea.strip().split('|')
                    if len(partes) >= 3:
                        cotizacion = partes[0].replace('SOL_PRC_', '')
                        numero_credito = partes[1]
                        fecha_raw = partes[2]
                        try:
                            fecha = datetime.strptime(fecha_raw, '%Y%m%d').date()
                        except:
                            fecha = None
                        registros.append({
                            'Cotizacion_Num': cotizacion,
                            'Número de Crédito 1': numero_credito,
                            'Fecha de Desembolso 1': fecha
                        })
    return pd.DataFrame(registros)

# ------------------------------------------------------
def unir_y_poblar(df_pricing, df_suite, df_txt):
    # Limpiar cotización
    df_pricing = limpiar_cotizaciones(df_pricing)
    df_pricing['Cotizacion_Num'] = df_pricing['Cotizacion_Num'].astype(str)

    # Suite Digital
    df_suite = df_suite.rename(columns={'x': 'Cotizacion_Num'})
    df_suite['Cotizacion_Num'] = df_suite['Cotizacion_Num'].astype(str)

    df_suite_min = df_suite[['Cotizacion_Num', 'NUMERO_CREDITO', 'FECHA_ACTUALIZACION']].drop_duplicates()
    df_merge = pd.merge(df_pricing, df_suite_min, on='Cotizacion_Num', how='left')

    # Poblar si está vacío
    df_merge['Número de Crédito 1'] = df_merge['Número de Crédito 1'].fillna(df_merge['NUMERO_CREDITO'])
    df_merge['Fecha de Desembolso 1'] = df_merge['Fecha de Desembolso 1'].fillna(df_merge['FECHA_ACTUALIZACION'])

    # Poblar desde TXT
    df_txt['Cotizacion_Num'] = df_txt['Cotizacion_Num'].astype(str)
    df_txt_min = df_txt[['Cotizacion_Num', 'Número de Crédito 1', 'Fecha de Desembolso 1']].drop_duplicates()

    df_final = pd.merge(
        df_merge.drop(columns=['NUMERO_CREDITO', 'FECHA_ACTUALIZACION'], errors='ignore'),
        df_txt_min,
        on='Cotizacion_Num',
        how='left',
        suffixes=('', '_txt')
    )

    df_final['Número de Crédito 1'] = df_final['Número de Crédito 1'].fillna(df_final['Número de Crédito 1_txt'])
    df_final['Fecha de Desembolso 1'] = df_final['Fecha de Desembolso 1'].fillna(df_final['Fecha de Desembolso 1_txt'])

    # ------------------------------------------------------
    # 🔥 FILTROS
    df_final = df_final[~df_final['Radicador'].str.upper().eq('MARIZA@BANCODEBOGOTA.COM.CO')]
    df_final = df_final[~df_final['Estado'].str.upper().eq('BORRADOR')]

    # ------------------------------------------------------
    # 🔥 RESOLUCIÓN INTELIGENTE DE DUPLICADOS (prioriza aprobados)
    prioridad_estado = {
        "APROBADO": 1,
        "APROBADA": 1,
        "APROBADA DIGITAL": 1,
        "APROBADA WEB": 1
    }

    df_final["prioridad"] = df_final["Estado"].str.upper().map(prioridad_estado).fillna(2)

    df_final = df_final.sort_values(by=["Número Cotización", "prioridad"], ascending=[True, True])
    df_final = df_final.drop_duplicates(subset=["Número Cotización"], keep="first")

    # Normalizar Número de Crédito 1: ceros a la izquierda hasta 14 dígitos
    df_final['Número de Crédito 1'] = df_final['Número de Crédito 1'].apply(normalizar_numero_credito)

    # Orden final
    df_final = df_final.sort_values(by="Número Cotización", ascending=True)

    # Limpiar columnas extra
    df_final = df_final.drop(columns=[
        'Cotizacion_Num',
        'Número de Crédito 1_txt',
        'Fecha de Desembolso 1_txt',
        'prioridad'
    ], errors='ignore')

    return df_final

# ------------------------------------------------------
def guardar(df, salida=SALIDA_FINAL):
    df.to_excel(salida, index=False)
    print(f"✅ Reporte final guardado como: {salida}")

# ------------------------------------------------------
def _to_float(v):
    """Convierte a float valores con coma decimal."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    # convertir coma decimal a punto
    s = s.replace(",", ".")
    try:
        return float(s)
    except:
        return None

# ------------------------------------------------------
def pintar_columna_av_por_regla_spread():
    """
    Pinta SOLO la FUENTE de AV:
      - verde si Spread (AT) > Spread sugerido (AU)
      - rojo  si Spread (AT) < Spread sugerido (AU)
    """
    wb = load_workbook(SALIDA_FINAL)
    ws = wb.active

    verde = Color(rgb=COLOR_VERDE)
    rojo = Color(rgb=COLOR_ROJO)

    max_row = ws.max_row

    for r in range(2, max_row + 1):
        spread = _to_float(ws[f"{COL_SPREAD_REAL}{r}"].value)          # AT
        sugerido = _to_float(ws[f"{COL_SPREAD_SUGERIDO}{r}"].value)    # AU

        if spread is None or sugerido is None:
            continue

        celda_av = ws[f"{COL_DESVIACION}{r}"]  # AV
        fuente_actual = celda_av.font if celda_av.font else Font()
        nueva = copy(fuente_actual)

        if spread > sugerido:
            nueva.color = verde
            celda_av.font = nueva
        elif spread < sugerido:
            nueva.color = rojo
            celda_av.font = nueva
        # si son iguales, no se cambia

    wb.save(SALIDA_FINAL)
    print("✅ Columna AV pintada (fuente) usando AT vs AU.")

# ------------------------------------------------------
def main():
    df_pricing, df_suite = cargar_archivos_base()
    df_txt = procesar_archivos_txt(CARPETA_TXT)
    df_final = unir_y_poblar(df_pricing, df_suite, df_txt)
    guardar(df_final)

    # ✅ extra: pintar AV con la regla (sin recalcular desviación)
    pintar_columna_av_por_regla_spread()

if __name__ == "__main__":
    main()